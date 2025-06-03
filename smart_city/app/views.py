from .permissions import * 
from rest_framework.views import APIView
from rest_framework.generics import *
from rest_framework.response import Response
from rest_framework import status
from .serializers import *
from .models import * 
from openpyxl import load_workbook, Workbook
from django.conf import settings
import os
from rest_framework_simplejwt.views import TokenObtainPairView
from django.http import HttpResponse
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from .filters import SensorFIlter, HistoricoFilter, AmbienteFilter
from rest_framework.permissions import IsAuthenticated

def parse_float(val):
    if isinstance(val, str):
        val = val.replace(',', '.')
    try:
        return float(val)
    except:
        return None
    
class AmbienteView(APIView):
    http_method_names = ['get', 'post', 'put', 'delete']
    permission_classes = [IsDirectorOrOnlyRead]

    @swagger_auto_schema(
        operation_description='Show all ambientes and specific using <int:pk>',
        responses={
            200: AmbienteSerializer,
            400: 'Bad Request',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )
    
    def get(self, request, *args, pk=None):
        if pk: 
            try:
                ambientes = Ambientes.objects.get(pk=pk)
                serializer = AmbienteSerializer(ambientes)
                return Response(serializer.data, status=status.HTTP_200_OK)
            except Http404:
                raise Http404('Not Found')
            
        else:
            # Aplica o filtro
            filterset = AmbienteFilter(request.GET, queryset=Ambientes.objects.all())
            if not filterset.is_valid():
                return Response(filterset.errors, status=status.HTTP_400_BAD_REQUEST)
            
            # Obtém a queryset filtrada
            ambientes = filterset.qs
            serializer = AmbienteSerializer(ambientes, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
       
        
    swagger_auto_schema(
        operation_description='Create a ambiente', 
        request_body=AmbienteSerializer,
        responses={
            201: openapi.Response("Successfully created Ambiente ✅", AmbienteSerializer),
            400: "Bad Request"
        }
    )
    def post(self, request, *args, **kwargs):
        serializer = AmbienteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message':'Successfully created Ambiente ✅'}, status=status.HTTP_201_CREATED)
        return Response({'message':'Error to create Ambiente 😥'}, status=status.HTTP_400_BAD_REQUEST)

        
    @swagger_auto_schema(
        operation_description='Update ambiente by ID',
        request_body=AmbienteSerializer,
        responses={
            200: AmbienteSerializer,
            400: 'Bad Request',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )
    def put(self, request, pk=None): # <-- For the update 
        try: 
            ambiente = Ambientes.objects.get(pk=pk)
            serializer = AmbienteSerializer(ambiente, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'message': 'Successfully ✅✨', 'data':serializer.data}, status=status.HTTP_200_OK)
            return Response({'message':'Error processing request', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        except Http404:
            raise Http404('Not Found')
    
    @swagger_auto_schema(
        operation_description='Delete ambiente by ID',
        responses={
            204: 'Successfully deleted',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )
    def delete(self, request, pk=None):
        try:
            ambientes = Ambientes.objects.get(pk=pk)
            ambientes.delete()
            return Response({'message':'Successfully ✅'}, status=status.HTTP_204_NO_CONTENT)
        except Http404:
            raise Http404('Not Found')   

# View to import data into Excel to Mysql.    
class ImportAmbienteData(APIView):
    permission_classes = [IsDirectorOrOnlyRead]
    @swagger_auto_schema(auto_schema=None)
    def post(self, request):
        caminho_Arquivo = os.path.join(settings.BASE_DIR, '..', 'Dados Integrador', 'ambientesComId.xlsx')
        try:
            wb = load_workbook(caminho_Arquivo, data_only=True)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                sig, descricao, ni, responsavel = row[:4] # Variables in portugues to avoid mistakes

                sensor_data = {
                        'sig': sig,
                        'descricao':descricao,
                        'ni':ni, 
                        'responsavel':responsavel
                    }

                serializer = AmbienteSerializer(data=sensor_data)
                if serializer.is_valid():
                    serializer.save()
            return Response({'message': 'Data ''Ambiente'' imported successfully ✅'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    
        
class SensoresView(APIView): 
    http_method_names = ['get', 'post', 'put', 'delete']
    permission_classes = [IsDirectorOrOnlyRead]
    @swagger_auto_schema(
        operation_description='Show all Sensores and specific using <int:pk>',
        responses={
            200: SensoresSerializer,
            400: 'Bad Request',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )
    def get(self, request, *args, pk=None):

        if pk:
            try: 
                sensores = Sensores.objects.all()
                serializer = SensoresSerializer(sensores, many=True)
                return Response(serializer.data, status=status.HTTP_200_OK)
            except Http404:
                raise Http404('Not Found')
        else:
            # Aplica o filtro
            filterset = SensorFIlter(request.GET, queryset=Sensores.objects.all())
            if not filterset.is_valid():
                return Response(filterset.errors, status=status.HTTP_400_BAD_REQUEST)
            
            # Obtém a queryset filtrada
            sensores = filterset.qs
            serializer = SensoresSerializer(sensores, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
    @swagger_auto_schema(
        operation_description='Create a Sensor', 
        request_body=SensoresSerializer,
        responses={
            201: openapi.Response("Successfully created ✅", AmbienteSerializer),
            400: "Bad Request"
        }
    )
    def post(self, request, *args, **kwargs):
        serializer = SensoresSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message':'Successfully created ✅'}, status=status.HTTP_201_CREATED)
        return Response({'message':'Error to create ❌'}, status=status.HTTP_400_BAD_REQUEST)
    @swagger_auto_schema(
        operation_description='Update Sensores by ID',
        request_body=SensoresSerializer,
        responses={
            200: SensoresSerializer,
            400: 'Bad Request',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )
    def put(self, request, pk=None):
        try: 
            sensores = Sensores.objects.get(pk=pk)
            serializer = SensoresSerializer(sensores, data=request.data, partial=True)
            if serializer.is_valid():
                return Response({'message':'Successfully updated ✅', 'data': serializer.data}, status=status.HTTP_201_CREATED)
            return Response({'message':'Error in Request ❌', 'Error': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        except Http404:
            raise ('Not Found')
    @swagger_auto_schema(
        operation_description='Delete Sensores by ID',
        responses={
            204: 'Successfully deleted',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )   
    def delete(self, request, pk=None):
        try: 
            sensores = Sensores.objects.get(pk=pk)
            sensores.delete()
            return Response({'message':'Successfully deleted'}, status=status.HTTP_204_NO_CONTENT)
        except Http404:
            raise Http404('Not Found')
# View to import data into Excel to Mysql.
        
# Save 'Contador' data 
class ImportContador(APIView):
    # permission_classes =[IsDirector]
    def post(self, request):
        caminho_Arquivo = os.path.join(settings.BASE_DIR, '..', 'Dados Integrador', 'contador.xlsx')

        try:
            wb = load_workbook(caminho_Arquivo, data_only=True)
            sheet = wb.active

           
            for linha_planilha, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):               
                if not any(row):
                    continue

                sensor, mac, unidade, lat, lon, status_valor = row

                sensor_data = {
                        'sensor': sensor,
                        'mac_address': mac,
                        'unidade_medida': unidade,
                        'latitude': parse_float(lat),
                        'longitude': parse_float(lon),
                        'status': status_valor,
                    }

                serializer = SensoresSerializer(data=sensor_data)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response({
                    'mensagem': f'Erro na linha {linha_planilha}',
                    'erros': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({'mensagem': 'Importação concluída'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)  

# Save 'Luminosidade' data 
class ImportLuminosidade(APIView):
    # permission_classes =[IsDirector]
    def post(self, request):
        caminho_arquivo = os.path.join(settings.BASE_DIR, '..', 'Dados Integrador', 'luminosidade.xlsx')

        try:
            wb = load_workbook(caminho_arquivo, data_only=True)
            sheet = wb.active

           
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                sensor, mac, unidade, lat, lon, status_valor = row

                sensor_data = {
                        'sensor': sensor,
                        'mac_address': mac,
                        'unidade_medida': unidade,
                        'latitude': parse_float(lat),
                        'longitude': parse_float(lon),
                        'status': status_valor,
                    }

                serializer = SensoresSerializer(data=sensor_data)
                if serializer.is_valid():
                    serializer.save()
            return Response({'mensagem': 'Importação concluída'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# Save 'Temperatura'  
class ImportTemperatura(APIView):
    # permission_classes =[IsDirector]
    def post(self, request):
        caminho_arquivo = os.path.join(settings.BASE_DIR, '..', 'Dados Integrador', 'temperatura.xlsx')

        try:
            wb = load_workbook(caminho_arquivo, data_only=True)
            sheet = wb.active

           
            for linha_planilha, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                sensor, mac, unidade, lat, lon, status_valor = row

                sensor_data = {
                        'sensor': sensor,
                        'mac_address': mac,
                        'unidade_medida': unidade,
                        'latitude': parse_float(lat),
                        'longitude': parse_float(lon),
                        'status': status_valor,
                    }

                serializer = SensoresSerializer(data=sensor_data)
                if serializer.is_valid():
                    serializer.save()
            return Response({'mensagem': 'Importação concluída'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)      
# Save 'Umidade' 
class ImportUmidade(APIView):
    # permission_classes =[IsDirector]
    def post(self, request):
        caminho_arquivo = os.path.join(settings.BASE_DIR, '..', 'Dados Integrador', 'umidade.xlsx')

        try:
            wb = load_workbook(caminho_arquivo, data_only=True)
            sheet = wb.active

           
            for linha_planilha, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                sensor, mac, unidade, lat, lon, status_valor  = row

                sensor_data = {
                        'sensor': sensor,
                        'mac_address': mac,
                        'unidade_medida': unidade,
                        'latitude': parse_float(lat),
                        'longitude': parse_float(lon),
                        'status': status_valor,
                    }

                serializer = SensoresSerializer(data=sensor_data)
                if serializer.is_valid():
                    serializer.save()
            return Response({'mensagem': 'Importação concluída'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)  
           

class HistoricoView(APIView):
    http_method_names = ['get', 'post', 'put', 'delete']
    permission_classes = [IsDirectorOrOnlyRead]
    @swagger_auto_schema(
        operation_description='Show all Historico and specific using <int:pk>',
        responses={
            200: HistoricoSerializer,
            400: 'Bad Request',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )
    def get(self, request, *args, pk=None):

        if pk:
            try: 
                historico = Historico.objects.all()
                serializer = HistoricoSerializer(historico, many=True)
                return Response(serializer.data, status=status.HTTP_200_OK)
            except Http404:
                raise Http404('Not Found')
        else:
            # Aplica o filtro
            filterset = HistoricoFilter(request.GET, queryset=Historico.objects.all())
            if not filterset.is_valid():
                return Response(filterset.errors, status=status.HTTP_400_BAD_REQUEST)
            
            # Obtém a queryset filtrada
            historico = filterset.qs
            serializer = HistoricoSerializer(historico, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
    @swagger_auto_schema(
        operation_description='Create a Historico', 
        request_body=HistoricoSerializer,
        responses={
            201: openapi.Response("Successfully created ✅", HistoricoSerializer),
            400: "Bad Request"
        }
    )
    def post(self, request, *args, **kwargs):
        serializer = HistoricoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message':'Successfully created ✅'}, status=status.HTTP_201_CREATED)
        return Response({'message':'Error to create ❌'}, status=status.HTTP_400_BAD_REQUEST)
    @swagger_auto_schema(
        operation_description='Update Historico by ID',
        request_body=HistoricoSerializer,
        responses={
            200: HistoricoSerializer, 
            400: 'Bad Request',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    )
    def put(self, request, pk=None):
        try: 
            historico = Historico.objects.get(pk=pk)
            serializer = HistoricoSerializer(historico, data=request.data, partial=True)
            if serializer.is_valid():
                return Response({'message':'Successfully updated ✅', 'data': serializer.data}, status=status.HTTP_201_CREATED)
            return Response({'message':'Error in Request ❌', 'Error': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        except Http404:
            raise ('Not Found')
    @swagger_auto_schema(
        operation_description='Delete Historico',
        responses={
            204: 'Successfully deleted',
            404: 'Not Found',
            500: 'Internal Server Error'
        }
    ) 
    def delete(self, request, pk=None):
        try: 
            historico = Historico.objects.get(pk=pk)
            historico.delete()
            return Response({'message':'Successfully deleted'}, status=status.HTTP_204_NO_CONTENT)
        except Http404:
            raise Http404('Not Found')
# View to import data into Excel to Mysql.
class ImportHistorico(APIView):
    permission_classes = [IsDirectorOrOnlyRead]
    @swagger_auto_schema(auto_schema=None)
    def post(self, request):
        caminho_Arquivo = os.path.join(settings.BASE_DIR, '..', 'Dados Integrador', 'historico.xlsx')
  
        try:
            wb = load_workbook(caminho_Arquivo, data_only=True)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                sensor_id, ambiente_id, value, timestamp = row

                dados = {
                    'sensor': sensor_id,      
                    'ambiente': ambiente_id,  
                    'valor': float(value),
                    'timestamp': timestamp  
                }

                serializer = HistoricoSerializer(data=dados)
                if serializer.is_valid():
                    serializer.save()
                else:
                    print(f"Serializer Error: {serializer.errors}")

            return Response({
                'message': 'Successfully imported 😃',
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class LoginView(TokenObtainPairView):
    serializer_class = LoginSerializer

class CreateUserView(APIView):
    permission_classes = [IsDirectorOrOnlyRead] 
    @swagger_auto_schema(
        operation_description="Cria um novo usuário. Apenas Diretores ou Administradores com permissão `add_usuario` podem realizar esta operação.",
        request_body=UsuarioSerializer,
        responses={
            201: openapi.Response(description="Usuário criado com sucesso"),
            400: "Dados inválidos",
            403: "Sem permissão"
        }
    )
    def post(self, request):
        if not request.user.has_perm('app.add_usuario'):  
            return Response({"detail": "Você não tem permissão para criar usuários."}, status=status.HTTP_403_FORBIDDEN)
        
        serializer = UsuarioSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()  
            return Response({
                'message':"usuario criado"
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(
        operation_description="Lista todos Usuarios",
        responses={
            200: UsuarioSerializer,
            404: 'Não Encontrado', 
            500: 'Erro na requisição', 
            }
    )
    
    def get(self, request, *args, **kwargs):
            return super().get(request, *args, **kwargs)
    
class UpdateDeleteDetailUsuario(RetrieveUpdateDestroyAPIView):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsDirectorOrOnlyRead]

    @swagger_auto_schema(
        operation_description='Retorna todos os usuarios',
        responses={
            200: UsuarioSerializer, 
            404: 'Não encontrado',  
            400: 'Erro na requisição', 
        }
    )
    def get(self, request, *args, **kwargs):
        try: 
            usuario = self.get_object()
            serializer = self.get_serializer(usuario)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Http404:
            raise Http404("Usuario não encontrando")
    @swagger_auto_schema(
            operation_description='Atualiza os usuarios', 
            request_body = UsuarioSerializer, 
            responses={
                201: UsuarioSerializer, 
                404: 'Não Encontrado', 
                500: 'Erro na requisição'
            }
    )
    def put(self, request, *args, **kwargs):
        try:
            usuario = self.get_object()
            serializer = self.get_serializer(usuario, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'message': 'Usuario atualizado com sucesso', 'data': serializer.data}, status=status.HTTP_200_OK)
            return Response({'message': 'Erro ao processar'}, status=status.HTTP_400_BAD_REQUEST)
        except Http404:
            raise Http404("Usuario não encontrando")
    
    @swagger_auto_schema(
        operation_description='Deleta o Usuario',
        responses={
            204: 'Deletado com sucesso!',  
            404: 'Não encontrado',  
            500:  'Erro na requisição'
        }
    )

    def delete(self, request, *args, **kwargs):
        try:
            usuario = self.get_object()
            serializer = self.get_serializer(usuario)
            return Response({'message': 'usuario apagado com sucesso'}, status=status.HTTP_204_NO_CONTENT)
        except Http404:
            raise Http404("Usuario não encontrando")

class SaveUser(APIView):
    # permission_classes = [IsDirector]
    def post(self, request):
        file_path = os.path.abspath(os.path.join(settings.BASE_DIR, '..', 'Dados Integrador', 'dadosUsuario.xlsx'))
        errors = []
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                try:
                    username, password, telefone, cargo = row

                    user_data = {
                        'username': username,
                        'password': password,
                        'telefone': telefone,
                        'cargo': cargo
                    }

                    serializer = UsuarioSerializer(data=user_data)
                    if serializer.is_valid():
                        serializer.save()
                    else:
                        errors.append({username: serializer.errors})
                except Exception as row_error:
                    errors.append({'row_error': str(row_error)})

            if errors:
                return Response({'message': 'Import completed with some errors ❌', 'errors': errors}, status=status.HTTP_207_MULTI_STATUS)
            return Response({'message': 'Usuários importados com sucesso ✅'}, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
######################################################################################################################################EXPORT 
class ExportSensores(APIView):
    permission_classes = [IsAuthenticated, IsDirectorOrOnlyRead]

    def get(self, request):
        sensores = Sensores.objects.all()

        wb = Workbook()

        ws_sensores = wb.active

        header = ['sensor','mac_address','unidade_medida','latitude','longitude','status']
        ws_sensores.append(header)

        for sensor in sensores:
            export_data = [
                sensor.sensor, 
                sensor.mac_address,
                sensor.unidade_medida, 
                sensor.latitude, 
                sensor.longitude, 
                sensor.status
            ]
            ws_sensores.append(export_data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sensores.xlsx"'
        wb.save(response)

        return response

class ExportAmbientes(APIView):
        permission_classes = [IsDirectorOrOnlyRead]

        def get(self, request):
            ambientes = Ambientes.objects.all()

            wb = Workbook()

            ws_ambientes = wb.active

            header = ['sig','descricao','ni','responsavel']
            ws_ambientes.append(header)

            for ambiente in ambientes:
                export_data = [
                    ambiente.sig,
                    ambiente.descricao, 
                    ambiente.ni,
                    ambiente.responsavel
                ]
                ws_ambientes.append(export_data)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="ambiente.xlsx"'
            wb.save(response)

            return response
class ExportHistorico(APIView):
    permission_classes = [IsDirectorOrOnlyRead]
    def get(self, request):
        historicos = Historico.objects.all()

        wb = Workbook()

        ws_historico = wb.active

        header = ['sensor',	'ambiente',	'valor', 'timestamp']
        ws_historico.append(header)

        for historico in historicos:
            export_data = [
                historico.sensor.sensor,
                historico.ambiente.descricao, 
                historico.valor,
                historico.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws_historico.append(export_data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="historico.xlsx"'
        wb.save(response)

        return response
    
